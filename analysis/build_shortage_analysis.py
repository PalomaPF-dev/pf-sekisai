# -*- coding: utf-8 -*-
"""積載計画・不足分析ワークブック生成スクリプト

5つの帳票エクスポート(①現状在庫一覧照会 / ②受払予定一覧照会 / ③投入一覧照会 /
④入荷実績一覧登録 / ⑤配船マスタ)を読み込み、実使用予定(②+③の結合)と
日次在庫推移・不足判定・配船別入荷・週別必要積載をまとめたExcelを出力する。

ファイル名は問わず、1行目のヘッダ内容で①〜⑤を自動判別する。
加工ルールの詳細は同ディレクトリの README.md を参照。

使い方:
    python3 build_shortage_analysis.py --input-dir ./exports --out 積載計画_不足分析.xlsx
"""
import argparse
import glob
import os
import re
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ヘッダ1行目に含まれる列名でファイルを判別する
SIGNATURES = {
    "inventory": {"在庫数量", "引当可能数量", "品目ＣＤ"},          # ①
    "planned":   {"受払予定伝票ＮＯ", "基準単位出庫数量"},          # ②
    "released":  {"投入残数量", "子品目ＣＤ", "投入予定日"},        # ③
    "arrivals":  {"入荷予定数量", "入荷連絡備考", "出荷実績ＮＯ"},   # ④
    "vessels":   {"shp_air_no", "wk_cd", "prt_dep_dt"},             # ⑤
}


def detect_files(input_dir):
    found = {}
    for path in sorted(glob.glob(os.path.join(input_dir, "*.xlsx"))):
        try:
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            hdr = {str(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if v}
            wb.close()
        except Exception:
            continue
        for key, sig in SIGNATURES.items():
            if sig <= hdr and key not in found:
                found[key] = path
                break
    missing = set(SIGNATURES) - set(found)
    if missing:
        raise SystemExit(f"入力ファイルが見つかりません: {sorted(missing)} (dir={input_dir})")
    return found


def parse_vessel_ref(ref):
    """④入荷連絡備考の配船番号を⑤配船マスタのキーに変換する。

    船便   : Y288-1K  -> (288, '1', '288-K')
    航空便 : YA289-7  -> (289, '2', '289-7')
    """
    if not isinstance(ref, str):
        return None
    m = re.match(r"^Y(A?)(\d{3})-(.+)$", ref.strip())
    if not m:
        return None
    is_air, wk, suffix = m.group(1) == "A", int(m.group(2)), m.group(3)
    if is_air:
        return wk, "2", f"{wk}-{suffix}"
    if suffix and suffix[0].isdigit():
        return wk, suffix[0], f"{wk}-{suffix[1:]}"
    return wk, "1", f"{wk}-{suffix}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", required=True, help="5つの帳票xlsxが入ったフォルダ")
    ap.add_argument("--out", default="積載計画_不足分析.xlsx")
    ap.add_argument("--as-of", default=None,
                    help="基準日 YYYY-MM-DD (省略時は②の最小予定日付の前日)")
    args = ap.parse_args()

    files = detect_files(args.input_dir)
    print("判別結果:", {k: os.path.basename(v) for k, v in files.items()}, file=sys.stderr)

    df1 = pd.read_excel(files["inventory"], dtype=str)
    df2 = pd.read_excel(files["planned"], dtype=str)
    df3 = pd.read_excel(files["released"], dtype=str)
    df4 = pd.read_excel(files["arrivals"], dtype=str)
    # ⑤配船マスタ(2行目は和名ヘッダのため除外)。論理削除行は除く
    df5 = pd.read_excel(files["vessels"], dtype=str, skiprows=[1])
    df5 = df5[df5["del_flg"] != "1"]
    vessel_master = {}
    for _, row in df5.iterrows():
        no = row.get("shp_air_no")
        if isinstance(no, str) and no.strip():
            vessel_master[no.strip()] = row

    # 対象品目 = 使用予定(②③)に登場する品目
    items = sorted(set(df2["品目ＣＤ"].dropna()) | set(df3["子品目ＣＤ"].dropna()))
    names = {}
    for cd, nm in df3[["子品目ＣＤ", "子品名"]].dropna().drop_duplicates().values:
        names[cd] = nm
    for cd, nm in df2[["品目ＣＤ", "品名"]].dropna().drop_duplicates().values:
        names.setdefault(cd, nm)

    d2_dates = pd.to_datetime(df2["予定日付"], errors="coerce")
    as_of = (pd.Timestamp(args.as_of) if args.as_of
             else (d2_dates.min() - pd.Timedelta(days=1)).normalize())
    print("基準日:", as_of.date(), " 対象品目:", len(items), file=sys.stderr)

    # ① 現在庫
    inv = df1[df1["品目ＣＤ"].isin(items)].copy()
    inv["qty"] = pd.to_numeric(inv["在庫数量"], errors="coerce")
    stock0 = inv.groupby("品目ＣＤ")["qty"].sum()

    # ③ 確定使用(投入残) — 過去日は基準日扱い
    d3 = df3[df3["子品目ＣＤ"].isin(items)].copy()
    d3["date"] = pd.to_datetime(d3["投入予定日"], errors="coerce").clip(lower=as_of)
    d3["qty"] = pd.to_numeric(d3["投入残数量"], errors="coerce")
    dem3 = d3.groupby(["子品目ＣＤ", "date"])["qty"].sum()

    # ② 未確定使用(出庫予定)
    d2 = df2[df2["品目ＣＤ"].isin(items)].copy()
    d2["date"] = d2_dates[d2.index]
    d2["qty"] = pd.to_numeric(d2["基準単位出庫数量"], errors="coerce")
    dem2 = d2.groupby(["品目ＣＤ", "date"])["qty"].sum()

    # ④ 入荷予定(未入荷)
    s4 = df4[df4["品目ＣＤ"].isin(items) & (df4["進捗状況区分名"] == "未入荷")].copy()
    s4["date"] = pd.to_datetime(s4["入荷予定日"], errors="coerce")
    s4["qty"] = pd.to_numeric(s4["入荷予定数量"], errors="coerce")
    sup = s4.groupby(["品目ＣＤ", "date"])["qty"].sum()
    sup_ship = s4.groupby(["入荷連絡備考", "品目ＣＤ"])["qty"].sum()
    ship_dates = s4.groupby("入荷連絡備考")["date"].agg(["min", "max"])
    sup_end = s4["date"].max()

    all_dates = sorted(set(dem3.index.get_level_values(1))
                       | set(dem2.index.get_level_values(1))
                       | set(sup.index.get_level_values(1)))
    if not all_dates:
        raise SystemExit("対象品目の予定データがありません")

    # ---------------- スタイル ----------------
    FONT = "Arial"
    f_base = Font(name=FONT, size=10)
    f_bold = Font(name=FONT, size=10, bold=True)
    f_title = Font(name=FONT, size=12, bold=True)
    f_note = Font(name=FONT, size=9, color="666666")
    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    f_hdr = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    fill_sub = PatternFill("solid", fgColor="DDEBF7")
    fill_warn = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM = "#,##0;[Red]-#,##0"
    DATE = "yyyy/mm/dd"

    def put(ws, r, c, v, font=f_base, fmt=None, fill=None, bd=True):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        if bd:
            cell.border = border
        return cell

    def header_row(ws, r, cols, start=1):
        for i, h in enumerate(cols):
            cell = put(ws, r, start + i, h, f_hdr, fill=fill_hdr)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def item_band(ws, row, ncols_per_item=3):
        for i, it in enumerate(items):
            put(ws, row, 2 + i * ncols_per_item, names.get(it, ""), f_bold, fill=fill_sub)
            ws.merge_cells(start_row=row, start_column=2 + i * ncols_per_item,
                           end_row=row, end_column=1 + (i + 1) * ncols_per_item)

    wb = Workbook()

    # ================= 実使用予定 =================
    ws_u = wb.active
    ws_u.title = "実使用予定"
    put(ws_u, 1, 1, "実使用予定（②受払予定 ＋ ③投入 の結合）", f_title, bd=False)
    put(ws_u, 2, 1, f"③=確定分(投入残数量・過去日は{as_of:%m/%d}扱い) ②=未確定分(基準単位出庫数量)。計=③+②", f_note, bd=False)
    hdr = ["日付"]
    for it in items:
        hdr += [f"{it}\n③確定", f"{it}\n②計画", f"{it}\n計"]
    header_row(ws_u, 4, hdr)
    item_band(ws_u, 3)
    r0 = 5
    for k, dt in enumerate(all_dates):
        r = r0 + k
        put(ws_u, r, 1, dt, fmt=DATE)
        for i, it in enumerate(items):
            c = 2 + i * 3
            put(ws_u, r, c, int(dem3.get((it, dt), 0) or 0), fmt=NUM)
            put(ws_u, r, c + 1, int(dem2.get((it, dt), 0) or 0), fmt=NUM)
            put(ws_u, r, c + 2, f"={get_column_letter(c)}{r}+{get_column_letter(c+1)}{r}",
                f_bold, fmt=NUM)
    last_u = r0 + len(all_dates) - 1
    r = last_u + 1
    put(ws_u, r, 1, "合計", f_bold, fill=fill_sub)
    for c in range(2, 2 + len(items) * 3):
        L = get_column_letter(c)
        put(ws_u, r, c, f"=SUM({L}{r0}:{L}{last_u})", f_bold, fmt=NUM, fill=fill_sub)
    ws_u.column_dimensions["A"].width = 12
    for c in range(2, 2 + len(items) * 3):
        ws_u.column_dimensions[get_column_letter(c)].width = 11
    ws_u.freeze_panes = "B5"

    # ================= 在庫推移 =================
    ws_b = wb.create_sheet("在庫推移")
    put(ws_b, 1, 1, "日次在庫推移と不足判定（追加の積載・入荷が無い場合）", f_title, bd=False)
    put(ws_b, 2, 1, "予定在庫 = 前日残 + 入荷予定(④未入荷) - 使用予定(実使用予定「計」)。マイナス=不足(赤)", f_note, bd=False)
    hdr = ["日付"]
    for it in items:
        hdr += [f"{it}\n入荷", f"{it}\n使用", f"{it}\n予定在庫"]
    header_row(ws_b, 4, hdr)
    item_band(ws_b, 3)
    r = 5
    put(ws_b, r, 1, f"現在庫({as_of:%m/%d})", f_bold, fill=fill_sub)
    for i, it in enumerate(items):
        put(ws_b, r, 2 + i * 3, "", fill=fill_sub)
        put(ws_b, r, 3 + i * 3, "", fill=fill_sub)
        put(ws_b, r, 4 + i * 3, int(stock0.get(it, 0)), f_bold, fmt=NUM, fill=fill_sub)
    r0b = 6
    for k, dt in enumerate(all_dates):
        r = r0b + k
        put(ws_b, r, 1, dt, fmt=DATE)
        ur = r0 + k
        for i, it in enumerate(items):
            c = 2 + i * 3
            put(ws_b, r, c, int(sup.get((it, dt), 0) or 0), fmt=NUM)
            put(ws_b, r, c + 1, f"=実使用予定!{get_column_letter(4+i*3)}{ur}", fmt=NUM)
            put(ws_b, r, c + 2,
                f"={get_column_letter(c+2)}{r-1}+{get_column_letter(c)}{r}-{get_column_letter(c+1)}{r}",
                f_bold, fmt=NUM)
    last_b = r0b + len(all_dates) - 1
    for i in range(len(items)):
        L = get_column_letter(4 + i * 3)
        ws_b.conditional_formatting.add(
            f"{L}{r0b}:{L}{last_b}",
            CellIsRule(operator="lessThan", formula=["0"], fill=fill_warn))
    ws_b.column_dimensions["A"].width = 13
    for c in range(2, 2 + len(items) * 3):
        ws_b.column_dimensions[get_column_letter(c)].width = 11
    ws_b.freeze_panes = "B6"

    # ================= 配船別入荷予定 =================
    ws_s = wb.create_sheet("配船別入荷予定")
    put(ws_s, 1, 1, "配船別 入荷予定（④の未入荷分・対象品目）", f_title, bd=False)
    put(ws_s, 2, 1, "配船番号=④入荷連絡備考。形式 Y{週}-{区分}{記号}(1=船便) / YA{週}-{連番}(航空便)。船名等は⑤配船マスタとの突合(未登録週は空欄)", f_note, bd=False)
    header_row(ws_s, 4, ["配船番号", "入荷予定日(最小)", "入荷予定日(最大)",
                         "船名(⑤)", "コンテナ(⑤)", "日本出港日(⑤)"]
               + [f"{it}\n{names.get(it, '')}" for it in items] + ["計"])
    ic0 = 7  # 品目列の開始
    r = 5
    unmatched = []
    for sh in sorted(ship_dates.index):
        put(ws_s, r, 1, sh, f_bold)
        put(ws_s, r, 2, ship_dates.loc[sh, "min"], fmt=DATE)
        put(ws_s, r, 3, ship_dates.loc[sh, "max"], fmt=DATE)
        parsed = parse_vessel_ref(sh)
        mrow = vessel_master.get(parsed[2]) if parsed else None
        if mrow is not None:
            put(ws_s, r, 4, mrow.get("shp_nm") or "")
            put(ws_s, r, 5, mrow.get("cntnr_sz") or "")
            dep = pd.to_datetime(mrow.get("prt_dep_dt"), errors="coerce")
            put(ws_s, r, 6, dep if pd.notna(dep) else "", fmt=DATE)
        else:
            unmatched.append(sh)
            for c in (4, 5, 6):
                put(ws_s, r, c, "")
        for i, it in enumerate(items):
            put(ws_s, r, ic0 + i, int(sup_ship.get((sh, it), 0) or 0), fmt=NUM)
        put(ws_s, r, ic0 + len(items),
            f"=SUM({get_column_letter(ic0)}{r}:{get_column_letter(ic0-1+len(items))}{r})",
            f_bold, fmt=NUM)
        r += 1
    put(ws_s, r, 1, "合計", f_bold, fill=fill_sub)
    for c in range(2, ic0):
        put(ws_s, r, c, "", fill=fill_sub)
    for i in range(len(items) + 1):
        L = get_column_letter(ic0 + i)
        put(ws_s, r, ic0 + i, f"=SUM({L}5:{L}{r-1})", f_bold, fmt=NUM, fill=fill_sub)
    if unmatched:
        put(ws_s, r + 2, 1,
            f"⑤配船マスタ未登録: {', '.join(unmatched)} (マスタ最新化後に再実行すると自動突合)", f_note, bd=False)
    for c, w in zip("ABCDEF", (12, 15, 15, 22, 12, 14)):
        ws_s.column_dimensions[c].width = w
    for i in range(len(items) + 1):
        ws_s.column_dimensions[get_column_letter(ic0 + i)].width = 13

    # ================= 積載計画案 =================
    # 週次バケットで「不足させないために各週に何個入荷(=積載)させるべきか」を算出。
    # 必要追加入荷 = MAX(0, 当週需要 + 安全在庫係数×翌週需要 - 前週末在庫 - 登録済入荷)
    ws_w = wb.create_sheet("積載計画案")
    put(ws_w, 1, 1, "積載計画案（週別の必要追加入荷量）", f_title, bd=False)
    fill_input = PatternFill("solid", fgColor="FFFF00")
    f_input = Font(name=FONT, size=10, bold=True, color="0000FF")
    put(ws_w, 2, 1, "安全在庫係数(翌週需要×)", f_bold, bd=False)
    put(ws_w, 2, 3, 0.5, f_input, fmt="0.00", fill=fill_input)
    put(ws_w, 2, 4, "海上リードタイム(週)", f_bold, bd=False)
    put(ws_w, 2, 6, 3, f_input, fmt="0", fill=fill_input)
    put(ws_w, 3, 1, "黄色セルは入力値(変更可)。必要追加入荷>0が積載対象。登録済入荷期間内(網掛け週)の不足は船便では間に合わないため航空便・前倒しを検討", f_note, bd=False)
    hdr = ["入荷週(月曜)", "配船番号(目安)", "積込目安(月曜)"]
    for it in items:
        hdr += [f"{it}\n需要", f"{it}\n登録済入荷", f"{it}\n必要追加入荷", f"{it}\n週末在庫"]
    header_row(ws_w, 5, hdr)
    for i, it in enumerate(items):
        put(ws_w, 4, 4 + i * 4, names.get(it, ""), f_bold, fill=fill_sub)
        ws_w.merge_cells(start_row=4, start_column=4 + i * 4,
                         end_row=4, end_column=7 + i * 4)

    wk_start = (min(all_dates) - pd.Timedelta(days=min(all_dates).weekday())).normalize()
    weeks = list(pd.date_range(wk_start, max(all_dates), freq="W-MON"))

    # 週→配船番号の対応(登録済は④の初回入荷週、以降はY番号を週次で外挿)
    week_vessels = {}
    max_ywk = 0
    for sh in ship_dates.index:
        parsed = parse_vessel_ref(sh)
        if parsed:
            max_ywk = max(max_ywk, parsed[0])
        wk = ship_dates.loc[sh, "min"]
        wk = (wk - pd.Timedelta(days=wk.weekday())).normalize()
        week_vessels.setdefault(wk, []).append(sh)
    last_reg_week = max(week_vessels) if week_vessels else wk_start

    r = 6
    put(ws_w, r, 1, "現在庫", f_bold, fill=fill_sub)
    put(ws_w, r, 2, "", fill=fill_sub)
    put(ws_w, r, 3, "", fill=fill_sub)
    for i in range(len(items)):
        for c in range(4 + i * 4, 7 + i * 4):
            put(ws_w, r, c, "", fill=fill_sub)
        put(ws_w, r, 7 + i * 4, f"=在庫推移!{get_column_letter(4+i*3)}5",
            f_bold, fmt=NUM, fill=fill_sub)
    r0w = 7
    fill_reg = PatternFill("solid", fgColor="F2F2F2")
    for k, wk in enumerate(weeks):
        r = r0w + k
        in_reg = wk <= last_reg_week
        row_fill = fill_reg if in_reg else None
        put(ws_w, r, 1, wk, fmt=DATE, fill=row_fill)
        if wk in week_vessels:
            label = ", ".join(sorted(week_vessels[wk]))
        elif wk > last_reg_week and max_ywk:
            label = f"Y{max_ywk + int((wk - last_reg_week).days // 7)}(推定)"
        else:
            label = ""
        put(ws_w, r, 2, label, fill=row_fill)
        put(ws_w, r, 3, f"=A{r}-$F$2*7", fmt=DATE, fill=row_fill)
        nxt = wk + pd.Timedelta(days=7)
        last_week = k == len(weeks) - 1
        for i in range(len(items)):
            c = 4 + i * 4
            Ld, Ls_, Lreq, Le = (get_column_letter(c), get_column_letter(c + 1),
                                 get_column_letter(c + 2), get_column_letter(c + 3))
            Lu = get_column_letter(4 + i * 3)  # 実使用予定「計」列
            Ls = get_column_letter(2 + i * 3)  # 在庫推移「入荷」列
            cond = (f'実使用予定!$A${r0}:$A${last_u},">="&$A{r},'
                    f'実使用予定!$A${r0}:$A${last_u},"<"&DATE({nxt.year},{nxt.month},{nxt.day})')
            put(ws_w, r, c, f"=SUMIFS(実使用予定!{Lu}{r0}:{Lu}{last_u},{cond})",
                fmt=NUM, fill=row_fill)
            cond2 = (f'在庫推移!$A${r0b}:$A${last_b},">="&$A{r},'
                     f'在庫推移!$A${r0b}:$A${last_b},"<"&DATE({nxt.year},{nxt.month},{nxt.day})')
            put(ws_w, r, c + 1, f"=SUMIFS(在庫推移!{Ls}{r0b}:{Ls}{last_b},{cond2})",
                fmt=NUM, fill=row_fill)
            nxt_dem = "0" if last_week else f"{Ld}{r+1}"
            put(ws_w, r, c + 2,
                f"=ROUNDUP(MAX(0,{Ld}{r}+$C$2*{nxt_dem}-{Le}{r-1}-{Ls_}{r}),0)",
                f_bold, fmt=NUM, fill=row_fill)
            put(ws_w, r, c + 3,
                f"={Le}{r-1}+{Ls_}{r}+{Lreq}{r}-{Ld}{r}", fmt=NUM, fill=row_fill)
    last_w = r0w + len(weeks) - 1
    r = last_w + 1
    put(ws_w, r, 1, "必要追加入荷 合計", f_bold, fill=fill_sub)
    put(ws_w, r, 2, "", fill=fill_sub)
    put(ws_w, r, 3, "", fill=fill_sub)
    for i in range(len(items)):
        for off in range(4):
            c = 4 + i * 4 + off
            if off == 2:
                L = get_column_letter(c)
                put(ws_w, r, c, f"=SUM({L}{r0w}:{L}{last_w})", f_bold, fmt=NUM, fill=fill_sub)
            else:
                put(ws_w, r, c, "", fill=fill_sub)
    for i in range(len(items)):
        Lreq = get_column_letter(6 + i * 4)
        ws_w.conditional_formatting.add(
            f"{Lreq}{r0w}:{Lreq}{last_w}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=fill_warn))
    ws_w.column_dimensions["A"].width = 12
    ws_w.column_dimensions["B"].width = 16
    ws_w.column_dimensions["C"].width = 12
    for c in range(4, 4 + len(items) * 4):
        ws_w.column_dimensions[get_column_letter(c)].width = 11
    ws_w.freeze_panes = "D6"

    # ================= サマリ =================
    ws_m = wb.create_sheet("サマリ", 0)
    put(ws_m, 1, 1, f"積載計画・不足分析サマリ（{as_of:%Y-%m-%d}時点）", f_title, bd=False)
    put(ws_m, 2, 1, "出典: ①現状在庫一覧照会・②受払予定一覧照会・③投入一覧照会・④入荷実績一覧登録・⑤配船マスタ", f_note, bd=False)
    sup_end_s = f"{sup_end:%m/%d}" if pd.notna(sup_end) else "-"
    header_row(ws_m, 4, ["品目ＣＤ", "品名", "現在庫", f"入荷予定計\n(〜{sup_end_s})",
                         f"使用予定計\n(〜{sup_end_s})", f"{sup_end_s}時点\n見込在庫",
                         "使用予定計\n(全期間)", "追加積載が無い場合\nの最終過不足",
                         "最初に在庫が\n不足する日"])
    r = 5
    for i, it in enumerate(items):
        put(ws_m, r, 1, it)
        put(ws_m, r, 2, names.get(it, ""))
        put(ws_m, r, 3, f"=在庫推移!{get_column_letter(4+i*3)}5", fmt=NUM)
        Ls = get_column_letter(2 + i * 3)
        Lu = get_column_letter(4 + i * 3)
        put(ws_m, r, 4, f"=SUM(在庫推移!{Ls}{r0b}:{Ls}{last_b})", fmt=NUM)
        put(ws_m, r, 5,
            f'=SUMIFS(実使用予定!{Lu}{r0}:{Lu}{last_u},実使用予定!$A${r0}:$A${last_u},'
            f'"<="&DATE({sup_end.year},{sup_end.month},{sup_end.day}))', fmt=NUM)
        put(ws_m, r, 6, f"=C{r}+D{r}-E{r}", f_bold, fmt=NUM)
        put(ws_m, r, 7, f"=実使用予定!{Lu}{last_u+1}", fmt=NUM)
        put(ws_m, r, 8, f"=C{r}+D{r}-G{r}", f_bold, fmt=NUM)
        Lb = get_column_letter(4 + i * 3)
        put(ws_m, r, 9,
            f"=IFERROR(INDEX(在庫推移!$A${r0b}:$A${last_b},"
            f"MATCH(TRUE,INDEX(在庫推移!{Lb}{r0b}:{Lb}{last_b}<0,0),0)),\"不足なし\")",
            f_bold, fmt=DATE)
        r += 1
    ws_m.conditional_formatting.add(
        f"F5:F{r-1}", CellIsRule(operator="lessThan", formula=["0"], fill=fill_warn))
    ws_m.conditional_formatting.add(
        f"H5:H{r-1}", CellIsRule(operator="lessThan", formula=["0"], fill=fill_warn))
    notes = [
        "",
        "【前提】",
        "・実使用予定 = ③投入一覧(確定分: 投入残数量) + ②受払予定一覧(未確定分: 基準単位出庫数量)。境界日は両方を合算。",
        "・③の過去日残(未消化)は基準日扱いで計上。",
        "・入荷予定 = ④入荷実績一覧登録の未入荷分。配船番号は「入荷連絡備考」欄。",
        f"・入荷予定の登録は{sup_end_s}まで。それ以降の使用予定は今後の配船への積載計画の対象。",
        "・在庫は①の在庫数量ベース(引当可能数量は未使用)。",
    ]
    rn = r + 1
    for t in notes:
        put(ws_m, rn, 1, t, f_note if not t.startswith("【") else f_bold, bd=False)
        rn += 1
    for i, w in enumerate([11, 20, 10, 12, 12, 12, 12, 16, 14]):
        ws_m.column_dimensions[get_column_letter(1 + i)].width = w
    for row in ws_m.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(args.out)
    print("saved:", args.out)


if __name__ == "__main__":
    main()
